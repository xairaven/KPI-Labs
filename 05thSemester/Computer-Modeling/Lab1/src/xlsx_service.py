import openpyxl


def get_table(filename):
    sheet = openpyxl.load_workbook(filename).active

    table = []
    for row in sheet.iter_rows(values_only=True):
        row_data = []

        # Process the data in the row
        for cell_value in row:
            row_data.append(cell_value)

        table.append(row_data)

    return table


def write_table(filename, model_values):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    for i, number in enumerate(model_values):
        worksheet.cell(row=i + 1, column=1, value=number)

    workbook.save(filename)
